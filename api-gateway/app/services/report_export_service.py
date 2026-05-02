"""CSV report export service for SSPA.

This module intentionally uses only the Python standard library plus a PostgreSQL
DB driver. It discovers the available reporting tables/views at runtime so it can
work with the current SSPA database even if the table is in public, analytics,
feature_store, or predictive schema.
"""

from __future__ import annotations

import csv
import io
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Iterable

try:  # Prefer psycopg2 because most FastAPI/Postgres stacks already use it.
    import psycopg2  # type: ignore
except ImportError:  # psycopg v3 fallback.
    import psycopg as psycopg2  # type: ignore


FEATURE_TABLE_CANDIDATES = [
    ("feature_store", "student_course_features"),
    ("feature_store", "student_course_features_vw"),
    ("analytics", "student_course_features"),
    ("public", "student_course_features"),
    ("public", "vw_student_course_features"),
    ("predictive", "student_course_predictions"),
    ("predictive", "student_risk_predictions"),
    ("public", "student_course_predictions"),
    ("public", "student_risk_predictions"),
]

INTERVENTION_TABLE_CANDIDATES = [
    ("public", "intervention_cases"),
    ("interventions", "intervention_cases"),
    ("intervention", "intervention_cases"),
    ("public", "interventions"),
]

STUDENT_COLUMNS = ["student_number", "student_id", "user_id", "eid"]
COURSE_COLUMNS = ["course_code", "site_id", "course_id", "site_title"]
SCORE_COLUMNS = ["avg_score", "average_score", "final_score", "score", "course_score"]
RISK_COLUMNS = ["risk_level", "risk_label", "predicted_risk", "risk", "risk_category"]
LECTURER_COLUMNS = ["lecturer_number", "lecturer_id", "instructor_id", "teacher_id"]


@dataclass(frozen=True)
class ReportSource:
    schema: str
    table: str
    columns: set[str]

    @property
    def relation(self) -> str:
        return f"{quote_ident(self.schema)}.{quote_ident(self.table)}"


def get_connection():
    """Create a PostgreSQL connection using common SSPA env names."""
    database_url = (
        os.getenv("DATABASE_URL")
        or os.getenv("SSPA_DATABASE_URL")
        or os.getenv("POSTGRES_URL")
    )

    if database_url:
        conn = psycopg2.connect(database_url)
    else:
        conn = psycopg2.connect(
            host=os.getenv("POSTGRES_HOST") or os.getenv("DB_HOST") or "localhost",
            port=os.getenv("POSTGRES_PORT") or os.getenv("DB_PORT") or "5432",
            dbname=(
                os.getenv("POSTGRES_DB")
                or os.getenv("DB_NAME")
                or "AnalyticalDataStore"
            ),
            user=(
                os.getenv("POSTGRES_USER")
                or os.getenv("DB_USER")
                or "SakaiLMSPlugin"
            ),
            password=(
                os.getenv("POSTGRES_PASSWORD")
                or os.getenv("DB_PASSWORD")
                or "SakaiLMSPlugin"
            ),
        )

    try:
        conn.autocommit = True
    except Exception:
        pass

    return conn


def quote_ident(value: str) -> str:
    """Quote a PostgreSQL identifier discovered from information_schema."""
    return '"' + str(value).replace('"', '""') + '"'


def first_existing(columns: set[str], candidates: Iterable[str]) -> str | None:
    for column in candidates:
        if column in columns:
            return column
    return None


def fetch_table_columns(conn, schema: str, table: str) -> set[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = %s
            """,
            (schema, table),
        )
        return {row[0] for row in cur.fetchall()}


def find_source(conn, candidates: list[tuple[str, str]], required_any: list[list[str]]) -> ReportSource | None:
    for schema, table in candidates:
        columns = fetch_table_columns(conn, schema, table)
        if not columns:
            continue

        valid = True
        for group in required_any:
            if not first_existing(columns, group):
                valid = False
                break

        if valid:
            return ReportSource(schema=schema, table=table, columns=columns)

    return None


def normalized_risk_expr(score_col: str | None, risk_col: str | None) -> str:
    risk_part = "NULL"
    if risk_col:
        rc = quote_ident(risk_col)
        risk_part = f"LOWER(REPLACE(CAST({rc} AS TEXT), '-', '_'))"

    score_part = "NULL"
    if score_col:
        score_part = f"NULLIF(CAST({quote_ident(score_col)} AS TEXT), '')::NUMERIC"

    return f"""
    CASE
      WHEN {risk_part} IN ('high', 'high_risk', 'at_risk', 'critical') THEN 'HIGH'
      WHEN {risk_part} IN ('moderate', 'medium', 'medium_risk') THEN 'MODERATE'
      WHEN {risk_part} IN ('on_track', 'low', 'low_risk', 'safe') THEN 'ON_TRACK'
      WHEN {score_part} IS NOT NULL AND {score_part} < 50 THEN 'HIGH'
      WHEN {score_part} IS NOT NULL AND {score_part} < 65 THEN 'MODERATE'
      WHEN {score_part} IS NOT NULL THEN 'ON_TRACK'
      ELSE 'UNKNOWN'
    END
    """


def optional_filter(source: ReportSource, column: str | None, param_name: str) -> str:
    if not column:
        return "TRUE"
    return f"(%({param_name})s IS NULL OR {quote_ident(column)} = %({param_name})s)"


def run_query(sql: str, params: dict[str, Any]) -> tuple[list[str], list[dict[str, Any]]]:
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            columns = [desc[0] for desc in cur.description]
            rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            return columns, rows


def export_at_risk_students(
    course_code: str | None = None,
    lecturer_number: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    with get_connection() as conn:
        source = find_source(
            conn,
            FEATURE_TABLE_CANDIDATES,
            [STUDENT_COLUMNS, COURSE_COLUMNS, SCORE_COLUMNS + RISK_COLUMNS],
        )

    if not source:
        return ["message"], [
            {
                "message": (
                    "No feature/prediction table found for at-risk export. "
                    "Run the feature-store or predictive pipeline first."
                )
            }
        ]

    student_col = first_existing(source.columns, STUDENT_COLUMNS)
    course_col = first_existing(source.columns, COURSE_COLUMNS)
    score_col = first_existing(source.columns, SCORE_COLUMNS)
    risk_col = first_existing(source.columns, RISK_COLUMNS)
    lecturer_col = first_existing(source.columns, LECTURER_COLUMNS)

    risk_expr = normalized_risk_expr(score_col, risk_col)
    score_select = f"{quote_ident(score_col)} AS avg_score" if score_col else "NULL AS avg_score"
    first_name = "first_name" if "first_name" in source.columns else None
    last_name = "last_name" if "last_name" in source.columns else None
    email = "email" if "email" in source.columns else None

    sql = f"""
    SELECT
      {quote_ident(student_col)} AS student_number,
      {quote_ident(course_col)} AS course_code,
      {f'{quote_ident(first_name)} AS first_name,' if first_name else "'' AS first_name,"}
      {f'{quote_ident(last_name)} AS last_name,' if last_name else "'' AS last_name,"}
      {f'{quote_ident(email)} AS email,' if email else "'' AS email,"}
      {score_select},
      {risk_expr} AS risk_level,
      %(exported_at)s AS exported_at
    FROM {source.relation}
    WHERE {optional_filter(source, course_col, 'course_code')}
      AND {optional_filter(source, lecturer_col, 'lecturer_number')}
      AND ({risk_expr}) IN ('HIGH', 'MODERATE')
    ORDER BY
      CASE ({risk_expr}) WHEN 'HIGH' THEN 1 WHEN 'MODERATE' THEN 2 ELSE 3 END,
      {quote_ident(course_col)},
      {quote_ident(student_col)}
    """

    return run_query(
        sql,
        {
            "course_code": course_code or None,
            "lecturer_number": lecturer_number or None,
            "exported_at": datetime.now(timezone.utc).isoformat(),
        },
    )


def export_course_summary(lecturer_number: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    with get_connection() as conn:
        source = find_source(
            conn,
            FEATURE_TABLE_CANDIDATES,
            [STUDENT_COLUMNS, COURSE_COLUMNS, SCORE_COLUMNS + RISK_COLUMNS],
        )

    if not source:
        return ["message"], [
            {
                "message": (
                    "No feature/prediction table found for course-summary export. "
                    "Run the feature-store or predictive pipeline first."
                )
            }
        ]

    student_col = first_existing(source.columns, STUDENT_COLUMNS)
    course_col = first_existing(source.columns, COURSE_COLUMNS)
    score_col = first_existing(source.columns, SCORE_COLUMNS)
    risk_col = first_existing(source.columns, RISK_COLUMNS)
    lecturer_col = first_existing(source.columns, LECTURER_COLUMNS)

    risk_expr = normalized_risk_expr(score_col, risk_col)
    avg_expr = (
        f"ROUND(AVG(NULLIF(CAST({quote_ident(score_col)} AS TEXT), '')::NUMERIC), 2)"
        if score_col
        else "NULL"
    )

    sql = f"""
    SELECT
      {quote_ident(course_col)} AS course_code,
      COUNT(DISTINCT {quote_ident(student_col)}) AS total_students,
      {avg_expr} AS average_score,
      SUM(CASE WHEN ({risk_expr}) = 'HIGH' THEN 1 ELSE 0 END) AS high_risk_students,
      SUM(CASE WHEN ({risk_expr}) = 'MODERATE' THEN 1 ELSE 0 END) AS moderate_students,
      SUM(CASE WHEN ({risk_expr}) = 'ON_TRACK' THEN 1 ELSE 0 END) AS on_track_students,
      SUM(CASE WHEN ({risk_expr}) = 'UNKNOWN' THEN 1 ELSE 0 END) AS unknown_students,
      %(exported_at)s AS exported_at
    FROM {source.relation}
    WHERE {optional_filter(source, lecturer_col, 'lecturer_number')}
    GROUP BY {quote_ident(course_col)}
    ORDER BY {quote_ident(course_col)}
    """

    return run_query(
        sql,
        {
            "lecturer_number": lecturer_number or None,
            "exported_at": datetime.now(timezone.utc).isoformat(),
        },
    )


def export_interventions(
    course_code: str | None = None,
    lecturer_number: str | None = None,
    status: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    with get_connection() as conn:
        source = find_source(
            conn,
            INTERVENTION_TABLE_CANDIDATES,
            [["status"], STUDENT_COLUMNS],
        )

    if not source:
        return ["message"], [
            {"message": "No intervention case table found for interventions export."}
        ]

    course_col = first_existing(source.columns, COURSE_COLUMNS)
    lecturer_col = first_existing(source.columns, [
        "created_by_identifier",
        "lecturer_number",
        "lecturer_id",
    ])
    status_col = "status" if "status" in source.columns else None

    preferred_columns = [
        "case_id",
        "student_number",
        "student_id",
        "course_code",
        "site_id",
        "risk_level",
        "reason",
        "priority",
        "status",
        "created_by_role",
        "created_by_identifier",
        "follow_up_date",
        "created_at",
        "updated_at",
        "note_count",
    ]
    selected = [column for column in preferred_columns if column in source.columns]
    if not selected:
        selected = sorted(source.columns)

    select_clause = ",\n      ".join(f"{quote_ident(column)}" for column in selected)

    sql = f"""
    SELECT
      {select_clause},
      %(exported_at)s AS exported_at
    FROM {source.relation}
    WHERE {optional_filter(source, course_col, 'course_code')}
      AND {optional_filter(source, lecturer_col, 'lecturer_number')}
      AND {optional_filter(source, status_col, 'status')}
    """

    return run_query(
        sql,
        {
            "course_code": course_code or None,
            "lecturer_number": lecturer_number or None,
            "status": status or None,
            "exported_at": datetime.now(timezone.utc).isoformat(),
        },
    )


def make_csv(rows: list[dict[str, Any]], headers: list[str]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()

    for row in rows:
        writer.writerow({key: row.get(key, "") for key in headers})

    return output.getvalue()


def make_excel(rows: list[dict[str, Any]], headers: list[str]) -> bytes:
    """Generate Excel format (.xlsx) report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Write headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_num, row in enumerate(rows, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = row.get(header, "")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in rows:
            if header in row:
                max_length = max(max_length, len(str(row[header] or "")))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(max_length + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def make_pdf(rows: list[dict[str, Any]], headers: list[str], title: str = "Report") -> bytes:
    """Generate PDF format report."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. "
            "Install it with: pip install reportlab"
        )

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch)
    elements = []

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.3*inch))

    # Prepare table data
    table_data = [headers]
    for row in rows:
        table_data.append([str(row.get(header, "")) for header in headers])

    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("366092")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("F0F0F0")]),
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def export_students_by_risk_level(
    risk_level: str,
    course_code: str | None = None,
    lecturer_number: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Export students filtered by specific risk level (HIGH, MODERATE, ON_TRACK)."""
    with get_connection() as conn:
        source = find_source(
            conn,
            FEATURE_TABLE_CANDIDATES,
            [STUDENT_COLUMNS, COURSE_COLUMNS, SCORE_COLUMNS + RISK_COLUMNS],
        )

    if not source:
        return ["message"], [
            {
                "message": (
                    "No feature/prediction table found. "
                    "Run the feature-store or predictive pipeline first."
                )
            }
        ]

    risk_level = risk_level.upper()
    if risk_level not in ("HIGH", "MODERATE", "ON_TRACK"):
        risk_level = "HIGH"

    student_col = first_existing(source.columns, STUDENT_COLUMNS)
    course_col = first_existing(source.columns, COURSE_COLUMNS)
    score_col = first_existing(source.columns, SCORE_COLUMNS)
    risk_col = first_existing(source.columns, RISK_COLUMNS)
    lecturer_col = first_existing(source.columns, LECTURER_COLUMNS)

    risk_expr = normalized_risk_expr(score_col, risk_col)
    score_select = f"{quote_ident(score_col)} AS avg_score" if score_col else "NULL AS avg_score"
    first_name = "first_name" if "first_name" in source.columns else None
    last_name = "last_name" if "last_name" in source.columns else None
    email = "email" if "email" in source.columns else None

    sql = f"""
    SELECT
      {quote_ident(student_col)} AS student_number,
      {quote_ident(course_col)} AS course_code,
      {f'{quote_ident(first_name)} AS first_name,' if first_name else "'' AS first_name,"}
      {f'{quote_ident(last_name)} AS last_name,' if last_name else "'' AS last_name,"}
      {f'{quote_ident(email)} AS email,' if email else "'' AS email,"}
      {score_select},
      {risk_expr} AS risk_level,
      %(exported_at)s AS exported_at
    FROM {source.relation}
    WHERE {optional_filter(source, course_col, 'course_code')}
      AND {optional_filter(source, lecturer_col, 'lecturer_number')}
      AND ({risk_expr}) = %(risk_level)s
    ORDER BY
      {quote_ident(course_col)},
      {quote_ident(student_col)}
    """

    return run_query(
        sql,
        {
            "course_code": course_code or None,
            "lecturer_number": lecturer_number or None,
            "risk_level": risk_level,
            "exported_at": datetime.now(timezone.utc).isoformat(),
        },
    )


def export_advisor_student_overview(
    advisor_id: str | None = None,
    risk_level: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Export student overview for advisors with all risk levels."""
    with get_connection() as conn:
        source = find_source(
            conn,
            FEATURE_TABLE_CANDIDATES,
            [STUDENT_COLUMNS, COURSE_COLUMNS, SCORE_COLUMNS + RISK_COLUMNS],
        )

    if not source:
        return ["message"], [
            {
                "message": (
                    "No feature/prediction table found. "
                    "Run the feature-store or predictive pipeline first."
                )
            }
        ]

    student_col = first_existing(source.columns, STUDENT_COLUMNS)
    course_col = first_existing(source.columns, COURSE_COLUMNS)
    score_col = first_existing(source.columns, SCORE_COLUMNS)
    risk_col = first_existing(source.columns, RISK_COLUMNS)

    risk_expr = normalized_risk_expr(score_col, risk_col)
    score_select = f"{quote_ident(score_col)} AS avg_score" if score_col else "NULL AS avg_score"
    first_name = "first_name" if "first_name" in source.columns else None
    last_name = "last_name" if "last_name" in source.columns else None
    email = "email" if "email" in source.columns else None

    risk_filter = ""
    params = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
    }

    if risk_level and risk_level.upper() in ("HIGH", "MODERATE", "ON_TRACK"):
        risk_filter = f"AND ({risk_expr}) = %s"
        params["risk_level"] = risk_level.upper()

    sql = f"""
    SELECT
      {quote_ident(student_col)} AS student_number,
      {quote_ident(course_col)} AS course_code,
      {f'{quote_ident(first_name)} AS first_name,' if first_name else "'' AS first_name,"}
      {f'{quote_ident(last_name)} AS last_name,' if last_name else "'' AS last_name,"}
      {f'{quote_ident(email)} AS email,' if email else "'' AS email,"}
      {score_select},
      {risk_expr} AS risk_level,
      %(exported_at)s AS exported_at
    FROM {source.relation}
    WHERE 1=1
    {risk_filter}
    ORDER BY
      CASE ({risk_expr}) WHEN 'HIGH' THEN 1 WHEN 'MODERATE' THEN 2 ELSE 3 END,
      {quote_ident(course_col)},
      {quote_ident(student_col)}
    """

    return run_query(sql, params)


def export_admin_institution_summary() -> tuple[list[str], list[dict[str, Any]]]:
    """Export institutional summary report for system administrators."""
    with get_connection() as conn:
        source = find_source(
            conn,
            FEATURE_TABLE_CANDIDATES,
            [STUDENT_COLUMNS, COURSE_COLUMNS, SCORE_COLUMNS + RISK_COLUMNS],
        )

    if not source:
        return ["message"], [
            {
                "message": (
                    "No feature/prediction table found. "
                    "Run the feature-store or predictive pipeline first."
                )
            }
        ]

    student_col = first_existing(source.columns, STUDENT_COLUMNS)
    course_col = first_existing(source.columns, COURSE_COLUMNS)
    score_col = first_existing(source.columns, SCORE_COLUMNS)
    risk_col = first_existing(source.columns, RISK_COLUMNS)

    risk_expr = normalized_risk_expr(score_col, risk_col)
    avg_expr = (
        f"ROUND(AVG(NULLIF(CAST({quote_ident(score_col)} AS TEXT), '')::NUMERIC), 2)"
        if score_col
        else "NULL"
    )

    sql = f"""
    SELECT
      {quote_ident(course_col)} AS course_code,
      COUNT(DISTINCT {quote_ident(student_col)}) AS total_students,
      {avg_expr} AS average_score,
      SUM(CASE WHEN ({risk_expr}) = 'HIGH' THEN 1 ELSE 0 END) AS high_risk_count,
      SUM(CASE WHEN ({risk_expr}) = 'MODERATE' THEN 1 ELSE 0 END) AS moderate_risk_count,
      SUM(CASE WHEN ({risk_expr}) = 'ON_TRACK' THEN 1 ELSE 0 END) AS on_track_count,
      ROUND(100.0 * SUM(CASE WHEN ({risk_expr}) = 'HIGH' THEN 1 ELSE 0 END) / 
            COUNT(DISTINCT {quote_ident(student_col)}), 2) AS high_risk_percentage,
      ROUND(100.0 * SUM(CASE WHEN ({risk_expr}) = 'MODERATE' THEN 1 ELSE 0 END) / 
            COUNT(DISTINCT {quote_ident(student_col)}), 2) AS moderate_risk_percentage,
      %(exported_at)s AS exported_at
    FROM {source.relation}
    GROUP BY {quote_ident(course_col)}
    ORDER BY 
      SUM(CASE WHEN ({risk_expr}) = 'HIGH' THEN 1 ELSE 0 END) DESC,
      {quote_ident(course_col)}
    """

    return run_query(
        sql,
        {
            "exported_at": datetime.now(timezone.utc).isoformat(),
        },
    )
